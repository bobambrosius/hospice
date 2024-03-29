from collections import Counter
from collections import namedtuple
from datetime import datetime
import re
from types import SimpleNamespace

from openpyxl import load_workbook

import const
import exceptions


class Person:
    """A Person is a human being.

    Attributes:
        name: (string)
            Full name of the person e.g. "John Doe".
        service: (string)
            Each shift needs a service 'verzorger' and a service 'algemeen'
            so we need to know wich type of service a person provides.
        shifts_per_weeks: (SimpleNamespace) shifts=int per_weeks=int
            On how many shifts in how many weeks the person
            wants to be scheduled.
        not_on_shifts_per_weekday: (dict)
            On which weekday on which shifts a person is not willing to work.
            key=(int) weekday, value = tuple of (int) shift
        not_in_timespan: (tuple)(date_object)[>(date_object)]
            On which days of the year quarter the person 
            doesn't want to be scheduled.
        preferred_shifts: (dict)
            Some volunteers prefer to be scheduled on a specific day and shift.
            key=(int) weekday, value = tuple of (int) shift
        availability_counter: (int)
            The number of times that the person 
            is available for scheduling per one or more weeks.
            The counter is initialised with shifts_per_week.shifts.
            The counter is reduced by 1 when a scheduling happens
            for the person.
            The counter is reset at the start of scheduling a new week 
            if the value is 0.
        weekend_counter: (int)
            Each person is obligated to run a shift in the weekend
            once a month.
            The counter has to be 4 for the person to be eligible for
            schediling in a weekend.
            When a person is scheduled in a weekend, the weekend_counter is
            reset to zero. At the start of scheduling a new week, 
            the counter is incremented by 1. While the counter is not yet 4,
            the person is not available for scheduling in the weekend.
    """
    def __init__(self):
        self.name = "" 
        self.service = ""
        self.shifts_per_weeks = ()
        self.not_on_shifts_per_weekday = dict()
        self.not_on_shifts_count = 0
        self.preferred_shifts = dict()
        self.not_in_timespan = ()
        self.availability_counter = 0 
        self.weekend_counter = 4

    def __repr__(self):
        return (
            f'{self.name}, '
            f'{self.service:10}, '
            f'shifts_per_weeks: ({self.shifts_per_weeks.shifts}, '
            f'{self.shifts_per_weeks.per_weeks}), '
            f'not_on_shifts_per_weekday: {self.not_on_shifts_per_weekday}, '
            f'preferred_shifts: {self.preferred_shifts}, '
            f'not_in_timespan: {self.not_in_timespan}, '
            f'availability_counter: {self.availability_counter}, '
            f'weekend_counter: {self.weekend_counter}'
        )


class Volunteers:
    """Volunteers is a collection of instances Person who work without
    fee for a hospice organisation.

    Attributes:
        persons: (tuple)
            all instances of Person.
        generalist_names: (set)
            A set of person names who's service is generalist.
        caretaker_names: (set)
            A set of person names who's service is caretaking.
    """
    def __init__(self, sourcefilename):

        self.sourcefilename = sourcefilename
        print(f'\nBestand lezen: "{self.sourcefilename}"...\n')

        # self.persons is a tuple with instances of class 'Person'
        self.persons = self._read_volunteersfile(self.sourcefilename)
        self._check_sanity("duplicate_names")

        # Get all 'generic' workers and all 'caretaker' workers.
        # Note: we use set operators on these groups and
        # the set operator 'difference' doesn't work on lists.
        # So we convert the list to a tuple and then to a set.
        self.generalist_names = set(tuple([
            p.name
            for p in self.persons
            if p.service == 'algemeen']))
        self.caretaker_names = set(tuple([ 
            p.name
            for p in self.persons
            if p.service == 'verzorger']))

    def search(self, namelist):
        """returns a list of instances of Person that have
        a matching name in namelist.
        """
        result = []
        for name in namelist:
            for person in self.persons:
                if person.name == name:
                    result.append(person)
        return result
    
    def get_optimal_person(self, namelist):
        """Return the person name who has the highest
        'not_on_shifts_per_weekday' shiftcount.
        If there is more than one person, return
        the last one found.
        Why do it? From the list of available persons
        for a shift, we prefer the person with most
        shifts unavailable. Other persons have a higher 
        availability, so we save them for a shift in 
        the future.
        """
        shiftcount = 0
        persons = [person for person in self.persons
                   if person.name in namelist]
        for person in persons:
            if (person.not_on_shifts_count >= shiftcount):
                # Choose the more or equal optimal person
                shiftcount = person.not_on_shifts_count
                personname = person.name
        return personname
            
    def show_count(self):
        """report how many persons of both service categories 
        are available this quarter.
        """
        print()
        print(f'Er zijn {len(self.caretaker_names)} verzorgers beschikbaar.')
        print(f'Er zijn {len(self.generalist_names)} algemenen beschikbaar.')
        print()

    def print_volunteers(self):
        print()
        for p in self.persons:
            print(p)

    def _day_and_shifts_to_dict(self,
            day_and_shifts_string, columnname, line_num):
        """The day_and_shifts_string is 
        for example 'ma:1, 2,3,4#  wo:3,4 # zo:4.'
        The function returns the dict: { 1: (1,2,3,4), 3: (3,4), 7: (4) }
        The weekddays are translated to isoweekday numbers,
        and the shifts are in a tuple.
        """
        if day_and_shifts_string:
            spaceless_string = day_and_shifts_string.replace(" ", "")
            
            # Check the input string
            self._check_sanity('day_and_shifts_string', spaceless_string,
                columnname=columnname, line_num=line_num)
            
            # Remove the last '#' AFTER the sanity check
            spaceless_string = spaceless_string.strip('#')
            
            # Make a list of items in the string 
            # with delimiter = '#':
            day_and_shifts_items = (i for i in spaceless_string.split('#'))
            # day_and_shift_list is e.g. ['ma:1,2,3,4', 'wo:3,4', 'zo:4'] 
            result_dict = {}
            for item in day_and_shifts_items:
                sep_weekday_and_shift = item.split(":")
                # The first sep_weekday_and_shift is ['ma', '1,2,3,4']
                # Now make a dict with key = isoweekday number
                # and value = tuple of shifts.
                key = const.WEEKDAY_LOOKUP[sep_weekday_and_shift[0]] 
                value = (int(i) for i in sep_weekday_and_shift[1].split(","))
                result_dict[key] = tuple(value)
            return result_dict
        else:
            return {}

    def _read_volunteersfile(self, sourcefile):
        """read a prepared .xls file <sourcefile>.
        The attributes are extracted from the column names.
        Read the values from the .xls file.
        Assign the values to the an instance of class 'Person'.
        Return a tuple of the instances 'Person'.
        """
        volunteers = []
        
        wb = load_workbook(filename=sourcefile, data_only=True)
        ws = wb.active
        min_col = 10
        max_col = min_col + 10  # Read eleven(!) columns
        reader = ws.iter_rows(min_col=min_col,
                max_col=max_col, values_only=True)
        # get names from column headers
        headers = next(reader)
        allowed_headers = ('Service', 'Achternaam', 'Voornaam', 'Tussenv', 
                'Actief', 'DienstenPerAantalWeken', 'NietOpDagEnDienst', 
                'VoorkeurDagEnDienst', 'NietInPeriode', 'VoorkeurTekst', 
                'NietSamenMet')
        for header in headers:
            if header not in allowed_headers:
                raise exceptions.SourceFileHeaderError(
                    f'Een of meer kolomkoppen komt niet voor '
                    f'in {allowed_headers}')
        try:
            Data = namedtuple("Data", headers)
        except ValueError as e:
            raise ValueError(f'De namen in de kolomkoppen mogen alleen '
                             f'alfanumerieke tekens bevatten. '
                             f'Kolomkop namen: {headers}') from e
        # start enumerating with line number 2
        for line_num, xls_data in enumerate(map(Data._make, reader), 2):
            # read only the Active persons
            if (xls_data.Actief):

                # Column Service
                # TODO Iemand kan zowel verzorger als algemeen zijn!
                # Moet dus een list worden i.p.v. string, 
                # met test op 'in' i.p.v. ==
                service = xls_data.Service or ""
                service = service.strip()
                self._check_sanity("service", service, "Service", line_num)

                # Columns Achternaam, Tussenv, Voornaam
                # Person name
                insert = xls_data.Tussenv or ""
                if insert.strip():
                    insert = " " + insert
                givenname = xls_data.Voornaam or ""
                surname = xls_data.Achternaam or "" 
                name = (givenname.strip() + insert + " " + surname.strip())

                # Column NietOpDagEnDienst
                not_on_shifts_per_weekday = (
                    xls_data.NietOpDagEnDienst or "")
                not_on_shifts_per_weekday = (
                    self._day_and_shifts_to_dict(not_on_shifts_per_weekday,
                    'NietOpDagEnDienst', line_num))
                self._check_sanity('day_and_shifts_dict',
                        not_on_shifts_per_weekday,
                        'NietOpDagEnDienst', line_num)
                
                # Count the number of not_in_shifts
                not_on_shifts_count = 0
                for weekday in not_on_shifts_per_weekday.keys():
                    not_on_shifts_count += len(
                        Counter(not_on_shifts_per_weekday[weekday]).values())
                pass
                
                # Column VoorkeurDagEnDienst
                # preferred_shifts (prefs)
                prefs_value = xls_data.VoorkeurDagEnDienst or ""
                pref_day_and_shifts = (
                    self._day_and_shifts_to_dict(prefs_value,
                    'VoorkeurDagEnDienst', line_num))
                self._check_sanity('day_and_shifts_dict',
                        pref_day_and_shifts,
                        'VoorkeurDagEnDienst', line_num)
                
                # Column DienstenPerAantalWeken
                # xls OpenOffice is confusing. Even though the column
                # is formatted as text, the value 1,1 is read as 
                # a float! After entering the value *again*
                # it is read as a string.
                shifts_per_week = xls_data.DienstenPerAantalWeken or ""
                shifts_per_week = shifts_per_week.replace(" ", "")
                self._check_sanity("shifts_per_weeks", 
                        shifts_per_week, "DienstenPerAantalWeken",
                        line_num)
                shifts_per_week = (shifts_per_week.split(","))
                shifts_per_weeks = SimpleNamespace(
                    shifts=int(shifts_per_week[0]),
                    per_weeks=int(shifts_per_week[1]))

                # availability_counter (no column)
                availability_counter = shifts_per_weeks.shifts
                
                # weekend counter (no column)
                # Initially everybody is available for weekends
                weekend_counter = const.WEEKENDCOUNTER

                # Column NietInPeriode
                not_in_timespan_value = xls_data.NietInPeriode or ""
                not_in_timespan = tuple( 
                    period for period in
                    not_in_timespan_value.replace(" ", "").split(",") 
                )
                self._check_sanity('dates_string', 
                    not_in_timespan, 'NietInPeriode', line_num)
                
                # Now we have all the data to instantiate a Person
                person = Person()
                person.name = name
                person.service = service
                person.not_on_shifts_per_weekday = (
                    not_on_shifts_per_weekday)
                person.not_on_shifts_count = not_on_shifts_count
                person.shifts_per_weeks = shifts_per_weeks
                person.not_in_timespan = not_in_timespan
                person.preferred_shifts = pref_day_and_shifts
                person.availability_counter = availability_counter
                person.weekend_counter = weekend_counter
                volunteers.append(person)
        return tuple(volunteers)

    def _check_sanity(self, test, operand=None, 
                      columnname=None, line_num=None):
        """Check the validity of the input data from the sourcefile.
        """
        # ------------------------------------------------------------
        # Start of helper functions
        def find_duplicate_personnames(nameslist, name):
            return [idx for idx, value in enumerate(nameslist)
                    if value == name]
        
        def check_day_and_shifts_count(operand):
            """Check for duplicate shifts in a day_and_shifts dict.
            """
            for weekday in operand.keys():
                cnt = Counter(operand[weekday])
                for shift in cnt.values():
                    if shift > 1:
                        return False
            return True
        
        def check_day_and_shifts_string(operand):
            # Return None if something is wrong.
            # return anything else if things are o.k.

            # There are no spaces in the operand.
            # operand example: di:1,2,3#wo:1,2,3#do:1,2,3#vr:1,2,3#
            
            # Check if the operand ends with a '#'.
            if operand[-1] != "#":
                return None

            # The pattern is
            # weekday + ':'
            #   + comma seperated shiftnumber e.g (1,2,3,4) or (1,2) or (3)
            # and then an endless repetition of
            #   '#' + (the first part, ending with '#').
            # Example: ma:1,2,3,4# di:1,2,3# zo:4#
            pattern = r'^(((ma|di|wo|do|vr|za|zo)[:][1-4]([,][1-4])*)[#])*$'
            return re.match(pattern, operand)

        def check_shifts_per_weeks(operand):
            # shifts_per_weeks must be like 1,2 or 3,2 or ...
            pattern = r'^(1,1|1,2|3,2|2,1|2,3)$'
            return re.match(pattern, operand)

        # End of helper functions
        # ------------------------------------------------------------
        match test:
            case 'duplicate_names': 
                nameslist = [p.name for p in self.persons]
                for name in nameslist:
                    if len(find_duplicate_personnames(nameslist, name)) > 1:
                        raise exceptions.DuplicatePersonnameError(name)
            
            case 'day_and_shifts_string':
                if not check_day_and_shifts_string(operand):
                    raise exceptions.DayAndShiftsStringError(
                        columnname, line_num, operand)
            
            case 'day_and_shifts_dict':
                if not check_day_and_shifts_count(operand):
                    raise exceptions.DayAndShiftsStringError(
                        columnname, line_num, operand)
            
            case 'shifts_per_weeks':
                if not check_shifts_per_weeks(operand):
                    raise exceptions.ShiftsPerWeeksError(
                        columnname, line_num, operand)
            
            case 'service':
                if not (operand in ('verzorger', 'algemeen')):
                    raise exceptions.ServicenameError(
                        columnname, line_num, operand)

            case 'dates_string':
                if all(operand):
                    # TODO gekopieerd uit init_agenda!
                    # TODO het jaar kan ook onjuist zijn, maar denk eraan
                    # dat de kwartalen over het jaar heen gaan.
                    for item in operand:
                        dates = item.split('>') 
                        try:
                            if len(dates) > 1:
                                startdate = datetime.strptime(dates[0], 
                                            const.DATEFORMAT).date()
                                enddate = datetime.strptime(dates[1], 
                                          const.DATEFORMAT).date()
                                if enddate < startdate:
                                    raise exceptions.DateTimespanError(
                                        columnname, line_num, item)
                            else:
                                # The variable is not important,
                                # only the execution of the function.
                                _ = datetime.strptime(dates[0], 
                                    const.DATEFORMAT).date()
                        except ValueError:
                            line_num = str(line_num)
                            raise exceptions.DateFormatError(
                                columnname, line_num, item)
            
            case _:
                raise exceptions.MissingCaseValueError(
                    'No match found for "test" in _check_sanity()')


if __name__ == '__main__':
    xls_filename = 'vrijwilligers-2023-kw2.xlsx'
    group = Volunteers(xls_filename)
    # group.print_volunteers()
    group.show_count()
